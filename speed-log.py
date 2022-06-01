#!/usr/bin/env python3

import datetime
import speedtest
import sys
import os
import openpyxl
from openpyxl.styles import Alignment

sptest = speedtest.Speedtest()
date_now = datetime.datetime.now()

def Get_Server() -> dict:
	return sptest.get_best_server() # as dictionary

def Test_Speed() -> dict:
	down = round(sptest.download() / 1000 / 1000) # Mbit/s
	up = round(sptest.upload() / 1000 / 1000) # Mbit/s

	return {"download": down,
			"upload": up,
			} # as dictionary in Mbit/s

def Excel_Save():
	print("Testing speed...")
	server = Get_Server()
	speed_result = Test_Speed()

	# Create initial excel file, if not exist
	path = os.path.dirname(__file__)+"\\"
	excel_file = path + "speed_history.xlsx"

	if not os.path.isfile(excel_file):
		wb = openpyxl.Workbook() # One time excel-file initializing
		ws_data = wb.active
		ws_data.title = "data"

		for i in range(1,7): # Merge and format first two row of each column
			ws_data.merge_cells(start_row=1, start_column=i, end_row=2, end_column=i)
			# Convert column 'number' [i] to letter [i:1 + ascii:96 = A]
			ws_data[chr(i + 96) + '1'].alignment = Alignment(horizontal="center", vertical="center", wrapText=True)

		ws_data['A1'] = "Date & Time"
		ws_data['B1'] = "Location"
		ws_data['C1'] = "Server"
		ws_data['D1'] = "Ping \nms"
		ws_data['E1'] = "Download \nMbit/s"
		ws_data['F1'] = "Upload \nMbit/s"

		try:
			wb.save(excel_file)
		except Exception as e:
			print(e)

	# Load excel-workbook and save new data
	try:
		wb = openpyxl.load_workbook(excel_file)
	except Exception as e:
		print(e)
	else:
		ws_data = wb["data"]
		last_row = str(ws_data.max_row + 1) # Find last row from data-worksheet

		ws_data["A" + last_row] = date_now
		ws_data["A" + last_row].number_format = 'dd.mm.yyyy h:mm'
		ws_data["B" + last_row] = server['country']
		ws_data["C" + last_row] = server['sponsor']
		ws_data["D" + last_row] = round(server['latency'])
		ws_data["E" + last_row] = speed_result['download']
		ws_data["F" + last_row] = speed_result['upload']

		try:
			wb.save(excel_file)
		except Exception as e:
			print(e)

def Show_Speed():
	server = Get_Server()
	speed_result = Test_Speed()
	print("Pvm:", date_now.strftime("%d.%m.%Y %H:%M"))
	print("Server:", server['country'] + ", " + server['sponsor'])
	print("Ping:", round(server['latency']), "ms")
	print("Download:", speed_result['download'], "Mbit/s")
	print("Upload:", speed_result['upload'], "Mbit/s")


argv = sys.argv

if "--show" in argv:
	Show_Speed()
	sys.exit()

Excel_Save()


"""
arg "-v" tallentaa ja tulostaa myös tulokset näytölle
"""
